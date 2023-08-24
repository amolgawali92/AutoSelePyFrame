import openpyxl

#File--->Workbook-->Sheets--->Rows--->Cells
file="C:\Drivers\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row  # count no.of rows in an Excel sheet
cols=sheet.max_column  # count no.of columns in an Excel sheet

# Reading all the rows & columns from Excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='    ')
    print()



