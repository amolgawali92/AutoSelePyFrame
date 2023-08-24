import openpyxl

#file="C:\\Drivers\\Book2.xlsx"

#workbook=openpyxl.load_workbook(file)
#sheet=workbook.active   # or sheet=workbook["data"]

# It's for same data add in to the Excel
#for r in range(1,6):
    #for c in range(1,4):
        #sheet.cell(r,c).value="Hello"

#workbook.save(file)

# It's for add Multiple data in to the Excel file

file="C:\\Drivers\\Book3.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active   # or sheet=workbook["data"]

sheet.cell(1,1).value=123
sheet.cell(1,2).value="A"
sheet.cell(1,3).value="QA"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="B"
sheet.cell(2,3).value="Sr.QA"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="C"
sheet.cell(3,3).value="Automation QA"

sheet.cell(4,1).value=101
sheet.cell(4,2).value="D"
sheet.cell(4,3).value="Sr.Automation QA"

workbook.save(file)   #save the file after entering the data.








