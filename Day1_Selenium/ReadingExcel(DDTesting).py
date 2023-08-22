#Data Driven Testing
import openpyxl

#File-->Workbook--->----Sheets--->Row---->Cells
file="D:\Selenium\DDTesting.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Data"]

rows=sheet.max_row  #count no.of rows in a excel sheet  6
cols=sheet.max_column  #count no.of columns in a excel sheet   4

#Reading all the rows & Columns fron excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='      ')
        print()
