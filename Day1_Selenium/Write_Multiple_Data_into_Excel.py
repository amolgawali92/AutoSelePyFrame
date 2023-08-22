#Write\Add Multiple data in to Excel...
import openpyxl   #We can work With Excel

file="D:\Selenium\DDTesting.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active      #(OR)..sheet=workbook["Data"]....get active sheet from excel

sheet.cell(1,1).value=1
sheet.cell(1,2).value="A"
sheet.cell(1,3).value="QA"
sheet.cell(1,4).value=1000000

sheet.cell(2,1).value=2
sheet.cell(2,2).value="B"
sheet.cell(2,3).value="Dev"
sheet.cell(2,4).value=400000

sheet.cell(3,1).value=3
sheet.cell(3,2).value="C"
sheet.cell(3,3).value="JS Developer"
sheet.cell(3,4).value=1200000

workbook.save(file)    #For save the File after Entering the Data...
